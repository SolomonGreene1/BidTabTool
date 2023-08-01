# https://python-textbok.readthedocs.io/en/1.0/Introduction_to_GUI_Programming.html
# https://likegeeks.com/python-gui-examples-tkinter-tutorial/
# Created By James Schroeder, edited by Solomon Greene

import tkinter as tk
from tkinter import Tk, Label, Button, Entry, scrolledtext, Checkbutton, END, filedialog, ttk
from ttkthemes import ThemedStyle
import tkinter.messagebox as messagebox
import xlrd
import glob
import os
import csv
from datetime import date
from PIL import ImageTk, Image
import time
import requests
import pandas as pd
from bs4 import BeautifulSoup
import re
import requests
import openpyxl as pxl
from openpyxl.styles import Font
import urllib.request
from tqdm import tqdm
import sys
import getpass
import subprocess

class CustomProgressBar:
	## This class used for printing progress updates to the window when downloading multiple files
    def __init__(self, total_files, txt):
        self.total_files = total_files
        self.current_file = 0
        self.txt = txt

    def update_progress(self):
        self.current_file += 1
        progress = self.current_file / self.total_files * 100
        self.txt.insert(tk.END, f"Progress: {progress:.2f}%\n") #progress bar line
        self.txt.see(tk.END)
        self.txt.update_idletasks()

class GUIwindow:
	
	def __init__(self, master):
		"Will be used to create and modify the GUI"
		## Initial setup
		self.master = master
		screen_width = master.winfo_screenwidth() #info for centering on user screen
		x_pos = (screen_width -1250)//2
		master.geometry(f'1250x750+{x_pos}+0')	#Height, width, xpos, ypos on screen
		master.title("Bid Tab Tool")
			
		## Expand text box to fill screen
		master.grid_rowconfigure(2, weight =1) #expands row 2 to fill screen
		master.grid_columnconfigure(1, weight =1) #expands column 1 to fill screen
		
		## Main Text box setup
		self.tbox = Entry(master, bd=5)
		self.tbox.grid(row=1, column=1, sticky = "w")
		self.tbox.focus()
		self.txt = scrolledtext.ScrolledText(master)
		self.txt.grid(row=2, column=1, columnspan=4, sticky = "wens")
		
		## Save as .csv checkbox setup
		self.chk_state = tk.BooleanVar()
		self.chk_state.set(True)
		#today = date.today() #get date for textbox label
		#date_str = today.strftime("%Y-%m-%d_Bid_Tabulations") 
		#self.chk = Checkbutton(master, text='Save as .CSV? (Saved in ' + date_str +")" , variable=self.chk_state)
		#self.chk.grid(row=1, column=3)
		
		## Auto scroll checkbox setup
		self.auto_scroll_var = tk.BooleanVar()
		self.auto_scroll_var.set(True)
		self.auto_scroll_checkbutton = Checkbutton(master, text ="Auto Scroll", variable = self.auto_scroll_var)
		self.auto_scroll_checkbutton.grid(row =3, column = 2)
		
		## Search by name checkbox setup
		self.by_name_var = tk.BooleanVar()
		self.by_name_var.set(False)
		self.by_name_checkbutton = Checkbutton(master, text ="Search for pay-item by name", variable = self.by_name_var)
		self.by_name_checkbutton.grid(row =2, column =0, sticky = "n")
		
		## Variables
		self.cancel_search = False # flag to indicate if search is cancelled
		self.last_search ="" # stores the last searched item for cancellation process
		self.auto_scroll = True # controls text auto scroll
		self.search_canceled = False # tracks search status
		self.csv_list = [] #stores names of all completed searches that saved as csv for cancellation delete
		self.bulk_search = False #keeps track of whether bulk search is implemented
		self.case1 = False
		self.case2 = False
		self.case3 = False # cases for save csv post search to delete correct amount of lines for formatting
		self.file_path_inserted = False
		self.abbreviated_names = [] #stores the abbreviated_names found in the contract for file naming purposes
		self.shortest_abbreviated = ""
		self.current_shortest_name = ""
		self.csv_file_path = "" #global csv file path
		self.bulk_num_items =0 #used so search label tracks what item you are on
		self.bulk_count = 0
		self.save_txt_item_found = False #flag to not save txt file even if user presses save for an empty entry
		self.to_download_list = [] #stores list of missing files to download
		current_script = os.path.abspath(sys.argv[0])
		current_dir = os.path.dirname(current_script)
		self.parent_dir = os.path.dirname(current_dir)
		self.BTT_Outputs_file_path = current_dir
		self.image_file_path = os.path.join(self.BTT_Outputs_file_path, "JWI Gray Logo.jpg")
		self.tollway_coded_file_path = os.path.join(self.BTT_Outputs_file_path,  "Tollway Coded Pay Items 11_07_2022.xlsx")
		self.IDOT_coded_file_path = os.path.join(self.BTT_Outputs_file_path, "20230804HWYCodedPayItems.xlsx")
		self.source_data_file_path = self.BTT_Outputs_file_path
		## Folder path info
		today = date.today() 
		date_str = today.strftime("%Y-%m-%d_Bid_Tabulations")
		parent_folder = "BTT Outputs"
		self.folder_path = os.path.join(self.parent_dir, parent_folder, date_str) #folder name
		
		## Iterate through folder, adding any pre-existing .csv files from the current date to the csv_list
		if os.path.exists(self.folder_path) and os.path.isdir(self.folder_path):
			for file_name in os.listdir(self.folder_path):
				#Check if csv
				if file_name.endswith(".csv"):
					#remove ending and add to csv_list
					csv_name = file_name.split("_",1)[0]
					self.csv_list.append(csv_name)
		
		## Keybindings
		master.bind('<Return>', self.entr) #for enter button
		self.txt.bind('<Button-1>', self.toggle_auto_scroll) #for click to stop scroll
		self.tbox.bind("<Button-3>", self.show_menu) #for right click copy paste in search box
		self.txt.bind("<Button-3>", self.show_txt_menu) #for right click copy paste in display data box
		
		## GUI Text Labels
		self.label = Label(master, text="Enter Pay Item Number: ")
		self.label.grid(row=1, column=0, sticky = "e")
		
		self.searchLabel = Label(master, font = ("Arial", 11, "bold"), fg = "green")
		self.searchLabel.grid(row=1, column=1, padx =180)
		
		self.saveLabel = Label(master)
		self.saveLabel.grid(row=3, column=3, pady =5 )
		
		## GUI Buttons
		self.search_button = Button(master, text="Search", command=self.search)
		self.search_button.grid(row=1, column=1, sticky = "w", padx =132)
		
		#self.save_txt_button = Button(master, text="Save as .txt", command=self.save_txt, state=tk.DISABLED)
		#self.save_txt_button.grid(row=3, column=1, sticky = "w")
		
		#self.save_csv_button = Button(master, text="Save as .csv", command=self.save_csv, state=tk.DISABLED)
		#self.save_csv_button.grid(row=3, column=1, sticky = "w", padx = 80)

		self.clear_button = Button(master, text="Clear All", command=self.clearAll, state=tk.DISABLED)
		self.clear_button.grid(row=3, column=4, padx =5, pady =5)
		
		self.close_button = Button(master, text="Close", command=master.quit)
		self.close_button.grid(row=3, column=5, padx = 25)
		
		self.cancel_button = Button(master, text = "Cancel Search", command=self.cancel_search_func, state = tk.DISABLED)
		self.cancel_button.grid(row =3, column = 3)
		
		self.bulk_label = Label(master, text="Bulk Processing: ")
		self.bulk_label.grid(row=1, column=2, sticky = "w")
		
		self.bulk_search_button = Button(master, text="Choose File", command=self.getFile)
		self.bulk_search_button.grid(row=1, column=2, padx =93)
		
		self.compare_lists_button = Button(master, text = "Check for new Bid Tabs", command = self.compare_lists)
		self.compare_lists_button.grid(row=2, column = 0, sticky = "s", padx = 5, pady = 5)
		
		## Image Display, scaled from 3900x2517, scaled down by 22
		self.image_label = Label(master)
		self.image_label.grid(row =2, column = 0, sticky = "w", padx = 10, pady =10)
		self.display_image(self.image_file_path, 177, 114)
		
	## Keybinding functionality methods
	def entr(self,event):
		"Allows hitting Enter key to provide the same user experiance as clicking search button"
		self.master.unbind('<Return>')
		self.search()
		
	def toggle_auto_scroll(self, event=None):
		"Turns the auto scroll off for user input click"
		self.auto_scroll = False
	
	## Search Bar copy paste
	def show_menu(self, event):
		"Displays the pop-up when right click is pressed IN SEARCH BAR"
		## setup and both labels with the methods they call
		menu = tk.Menu(self.master, tearoff = False)
		menu.add_command(label = "Paste (Ctrl + V)", command = self.paste)
		menu.add_command(label = "Copy (Ctrl + C)", command = self.copy)
		
		## functionality for copying a selected portion of text
		try:
			selected_text = self.get_selected_text()
			if selected_text:
				menu.entryconfig("Copy (Ctrl + C)", state=tk.NORMAL)
			else: #copy option disabled because nothing selected
				menu.entryconfig("Copy (Ctrl + C)", state=tk.DISABLED)
			menu.tk_popup(event.x_root, event.y_root)
		finally:
			menu.grab_release()
		
	def paste(self):
		"Paste function IN SEARCH BAR"
		self.tbox.event_generate("<<Paste>>")
	
	def copy(self):
		"Copy function in SEARCH BAR"
		selected_text = self.get_selected_text()
		if selected_text:
			self.master.clipboard_clear()
			self.master.clipboard_append(selected_text)
		
	def get_selected_text(self):
		"Accesses selected text for SEARCH BAR copy"
		if self.tbox.selection_present():
			return self.tbox.selection_get()
		return ""
	
	## Main text box copy paste
	def show_txt_menu(self, event):
		"Displays the pop-up when right click is pressed in MAIN TEXT BOX"
		menu = tk.Menu(self.master, tearoff = False)
		menu.add_command(label = "Paste (Ctrl + V)", command = self.txt_paste)
		menu.add_command(label = "Copy (Ctrl + C)", command = self.txt_copy)
		try:
			selected_text = self.get_txt_selected_text()
			if selected_text:
				menu.entryconfig("Copy (Ctrl + C)", state=tk.NORMAL)
			else:
				menu.entryconfig("Copy (Ctrl + C)", state=tk.DISABLED)
			
			menu.tk_popup(event.x_root, event.y_root)
		finally:
			menu.grab_release()
		
	def txt_paste(self):
		"Paste function MAIN TEXT BOX"
		self.txt.event_generate("<<Paste>>")
		
	def get_txt_selected_text(self):
		"Acquire selected text MAIN TEXT BOX"
		if self.txt.tag_ranges(tk.SEL):
			return self.txt.get(tk.SEL_FIRST, tk.SEL_LAST)
		return ""
	
	def txt_copy(self):
		"Copy MAIN TEXT BOX"
		selected_text = self.get_txt_selected_text()
		if selected_text:
			self.master.clipboard_clear()
			self.master.clipboard_append(selected_text)
	
	## Website bid tab updating methods
	def compare_lists(self):
		"This method compares the current dowloaded bid tabs with all bid tabs on the website to check if new files are available"
		# Website URL
		website_url = "https://www.illinoistollwaybidding.com/jobs/678/specs/bid-tabulations"

		# File path for downloaded files
		downloaded_path = self.source_data_file_path
		
		## Display Configurations
		self.bulk_search_button.config(state=tk.DISABLED)
		self.search_button.config(state = tk.DISABLED)
		
		# Get website content
		response = requests.get(website_url)
		soup = BeautifulSoup(response.content, "html.parser")
		web_content = soup.prettify()
		lines = web_content.split("\n") #split lines by new line
		
		## Website end: Lists for editing format of data
		website_list, int_list, final_web_list = [], [], []
		# Iterate over each line of the html file and check if it contains ".xls" or ".xlsx"
		for line in lines:
			if ".xls" in line or ".xlsx" in line:
				website_list.append(line)
		# formatting
		for item in website_list:
			code = item.split()[0]
			int_list.append(code)	
		final_web_list = [code.split(".", 1)[0] if "." in code else code for code in int_list]
		
		## Downloaded end: Get list of downloaded files
		downloaded_list, codes_list, final_dowloaded_codes, second_list = [], [], [], []
		for root, dirs, files in os.walk(downloaded_path):
			for file in files:
				if file.endswith(".xlsx") or file.endswith(".xls"):
					downloaded_list.append(file)
		# formatting			
		for item in downloaded_list:
			code = item.split()[0]
			code = item.split(".", 1)[0] if "." in item else item
			codes_list.append(code)
		for item in codes_list:
			code = item.split()[0]
			second_list.append(code)
		final_downloaded_codes = second_list
		
		# Find missing files
		missing_files = [file for file in final_web_list if file not in final_downloaded_codes]
		
		# Look through missing files and save actual name of file to to_download_list 
		for file_num in missing_files:
			for entry in website_list:
				if str(file_num) in entry:
					name = entry.lstrip()
					self.to_download_list.append(name)
					
		# Display missing files in a tkinter message box
		message = f"The following new files are available. Press OK to download them \n\n{', '.join(missing_files)}"
		if missing_files:
			root = Tk()
			root.withdraw()
			result = messagebox.showinfo("New Bid Tabs Available", message, icon = "info", type="okcancel") #option to download
			if result == 'ok':
				total_files = len(self.to_download_list)
				progress_bar = CustomProgressBar(total_files, self.txt)
				for title in self.to_download_list:
					self.download_file(title)
					#print to self.txt
					progress_bar.update_progress()
					root.update() #yield control to main loop so updates happen
			if result =="ok":
				messagebox.showinfo("Download Complete", "All files have been downloaded.")
				self.clear_output()
		else:
			messagebox.showinfo("No New Bid Tabs Available", "You are up to date!")
		self.bulk_search_button.config(state=tk.ACTIVE)
		self.search_button.config(state = tk.ACTIVE)
	
	def download_file(self, title):
		"This method is used to download a bid tab excel sheet from the tollway website"
		#Get URL and parse through it
		url = "https://www.illinoistollwaybidding.com/jobs/678/specs/bid-tabulations"
		response = requests.get(url)
		soup = BeautifulSoup(response.text, 'html.parser')

		download_url = None
		
		#find the download URL for the given title
		td_elements = soup.find_all('td')
		for index, td in enumerate(td_elements):
			if td.text.strip() == title:
				download_button = td_elements[index+2].find('button')
				if download_button:
					download_url = download_button.get('onclick')
					download_url = re.search(r"'(.*?)'", download_url).group(1)
					break
		
		#upon finding a url, download it to the right path
		if download_url:
			file_name = title  
			save_location = self.source_data_file_path
			save_path = os.path.join(save_location, file_name)

			full_download_url = 'https://www.illinoistollwaybidding.com' + download_url
			urllib.request.urlretrieve(full_download_url, save_path)
			time.sleep(1) #used for processing
			
			#Extract year from HTML to save in right folder
			current_element = download_button
			year_match = None
			while current_element:
				#parse through looking for a year 20xx 
				if hasattr(current_element, 'text') and 'Bid Tabulations' in current_element.text:
					year_match = re.search(r'20(\d{2})', current_element.text)
					if year_match:
						break
				current_element = current_element.parent
			
			if year_match:
				year = "20" + year_match.group(1)
				folder_name = f"{year} Bid Tabulations"
				folder_path = os.path.join(save_location, folder_name)
				
				if not os.path.exists(folder_path):
					os.makedirs(folder_path)
				
				#save new name and location
				new_save_path = os.path.join(folder_path, file_name)
				os.rename(save_path, new_save_path)
				time.sleep(1)
				#call excel stripping function to reduce file size
				if new_save_path.lower().endswith('.xls'):
					converted_path = self.convert_to_xlsx(new_save_path)
					stripped_file = self.strip_empty_cells(converted_path)
				else:
					stripped_file = self.strip_empty_cells(new_save_path)
	
	def convert_to_xlsx(self, filename):
		"This handles the edge case that a file is uploaded to the website in .xls (outdated) format"
		converted_filename = filename[:-4] + ".xlsx"
		workbook = pxl.Workbook()
		# Load the .xls file using xlrd
		xls_workbook = xlrd.open_workbook(filename)
		sheet_names = xls_workbook.sheet_names()

		# Iterate through each sheet in the .xls file and copy the data to the .xlsx file
		for sheet_name in sheet_names:
			sheet = workbook.create_sheet(title=sheet_name)
			xls_sheet = xls_workbook.sheet_by_name(sheet_name)

			# Copy the cell values to the new sheet
			for row in range(xls_sheet.nrows):
				for col in range(xls_sheet.ncols):
					cell_value = xls_sheet.cell_value(row, col)
					sheet.cell(row=row+1, column=col+1).value = cell_value

			# Set font style for all cells
			font = Font(name="Calibri", size=11)
			for row in sheet.iter_rows(min_row=1, max_row=xls_sheet.nrows, min_col=1, max_col=xls_sheet.ncols):
				for cell in row:
					cell.font = font

			# Adjust column widths
			for column in range(1, xls_sheet.ncols + 1):
				sheet.column_dimensions[pxl.utils.get_column_letter(column)].width = 12

		# Remove the default sheet created by openpyxl and save the converted file
		del workbook["Sheet"]
		workbook.save(converted_filename)
		
		#remove the .xls file
		os.remove(filename)

		return converted_filename

	## Excel Stripper methods
	def strip_empty_cells(self, filename):
		# Read the Excel file into a pandas DataFrame
		df = pd.read_excel(filename)

		# Drop rows and columns with all NaN (empty) values
		df.dropna(how='all', axis=0, inplace=True)
		df.dropna(how='all', axis=1, inplace=True)

		# Save the stripped DataFrame back to an Excel file
		directory = os.path.dirname(filename)
		new_filename = os.path.join(directory, f"{os.path.basename(filename)}")
		df.to_excel(new_filename, index=False)

		return new_filename
	
	## Button functionality methods
	def getFile(self):
		"Gets input file for bulk search method"
		self.sourceListPath = filedialog.askopenfilename(initialdir = os.path.dirname(__file__), filetypes = (("Excel files","*.csv;*.xls;*.xlsx"),("Text files","*.txt"),("all files","*.*")))
		self.tbox.insert(0,self.sourceListPath)
		self.file_path_inserted = True
		
	def bulk(self):
		"This method allows for the functionality of bulk searching multiple pay items at a time"
		## Open inputted sheet from getFile
		srcwb = xlrd.open_workbook(self.sourceListPath)
		srcsheet = srcwb.sheet_by_index(0)
		
		## Get number of filled rows with a pay item number
		self.bulk_num_items = sum(1 for row in range(srcsheet.nrows) if srcsheet.cell_value(row, 0))
		self.bulk_count = 1
		
		## Display Configurations
		self.compare_lists_button.config(state=tk.DISABLED)
		self.bulk_search_button.config(state=tk.DISABLED)
		#self.chk.config(state=tk.DISABLED)
		self.search_button.config(state=tk.DISABLED)
		#self.save_txt_button.config(state=tk.DISABLED)
		#self.save_csv_button.config(state=tk.DISABLED)
		self.clear_button.config(state=tk.DISABLED)
		self.search_button.config(state=tk.ACTIVE)
		search_label_string = "Searching! "+ str(self.bulk_count) + "/" + str(self.bulk_num_items)
		self.searchLabel.config(text=search_label_string)
		self.cancel_button.config(state=tk.ACTIVE)
		
		## For each row (PI)...
		for srci in range(srcsheet.nrows):
			if self.search_canceled:
				self.searchLabel.config(text="Search Canceled!")
				self.compare_lists_button.config(state=tk.ACTIVE)
				self.bulk_search_button.config(state=tk.ACTIVE)
				return
			self.bulk_count = self.bulk_count +1
			self.shortest_abbreviated = ""
			self.abbreviated_names.clear()
			self.getBids(srcsheet.cell_value(srci, 0))
			self.save_txt() #save as txt. Save as csv if checkbox is checked
			self.save_csv()
			self.clearAll_2() #Alternate method used so search label displays correctly
		self.searchLabel.config(text="Done!")
		self.cancel_button.config(state=tk.DISABLED)
		self.compare_lists_button.config(state=tk.ACTIVE)
		self.bulk_search_button.config(state=tk.ACTIVE)
		self.clear_button.config(state=tk.ACTIVE)
		#self.chk.config(state=tk.ACTIVE)
		
	def search_by_name(self):
		"This function pulls pay-item codes based on a user inputted name"
		## initial check if bulk search
		if self.bulk_search:
			return
		
		self.item_found = False #flag for finding at least one item
		self.search_canceled = False #initialize search canceled upon new search
		data = self.tbox.get()
		if len(data) < 3: #arbitrary param to eliminate long searches that potentially dont terminate
			self.txt.insert("1.0","Seach term must have at least 3 characters.")
			self.clear_button.config(state = tk.ACTIVE)
			return
		
		## Display configurations
		self.compare_lists_button.config(state=tk.DISABLED)
		self.search_button.config(state=tk.DISABLED)
		#self.save_txt_button.config(state=tk.DISABLED)
		#self.save_csv_button.config(state=tk.DISABLED)
		self.clear_button.config(state=tk.DISABLED)
		self.searchLabel.config(text="Searching Now!")
		self.txt.insert("1.0", "List of matched items and names:\n\n")
		
		# open sheet to search
		# workbook1 = xlrd.open_workbook(r"E:\OneDrive - J.A. Watts Inc\Tollway General\Bid Tabulations\Source Data DO NOT EDIT\Tollway Coded Pay Items 11_07_2022.xlsx") #file path 
		workbook1 = xlrd.open_workbook(self.tollway_coded_file_path) #file path 
		sheet = workbook1.sheet_by_name("Sheet1") #sheet name
		#iterate through all rows in column 3 of the sheet
		for row_idx in range(1, sheet.nrows):
			code = str(sheet.cell_value(row_idx, 2)) #codes in third column
			name = str(sheet.cell_value(row_idx, 3)) #names in fourth
			if data.lower() in name.lower():
				self.item_found = True
				self.txt.insert("end", code + "    " + name +"\n")
		#same thing but with second sheet
		workbook2 = xlrd.open_workbook(self.IDOT_coded_file_path)
		sheet2 = workbook2.sheet_by_name("Sheet1")
		for row_idx in range(sheet2.nrows):
			code = str(sheet2.cell_value(row_idx, 0))
			name = str(sheet2.cell_value(row_idx, 1))
			if data.lower() in name.lower():
				self.item_found = True
				self.txt.insert("end", code + "    " + name +"\n")
		
		if not self.item_found:
			self.txt.insert("end", "No items found with name: " + data)
		if self.item_found:
			self.by_name_var.set(False) #reset checkbox if item found, assumes user now wants to search a code
			
		## Display configurations
		self.searchLabel.config(text="Done!")
		self.clear_button.config(state=tk.ACTIVE)
		self.compare_lists_button.config(state=tk.ACTIVE)
	
	def search(self):
		"Pulls in data from text box, and runs process to search for entered item"
		self.clear_output()
		
		##Variable initializations
		self.auto_scroll = True #resets auto scroll when new search entered so click stop functionality works
		self.search_canceled = False #initialize search canceled upon new search
		self.cancel_search = False
		self.case1 = False
		self.case2= False
		self.case3 = False
		self.shortest_abbreviated = ""
		self.abbreviated_names.clear()
		
		## search by name checkbox sends to func: search_by_name()
		if self.by_name_var.get():
			self.search_by_name()
			return
		if self.file_path_inserted:
			self.file_path_inserted = False
			self.bulk_search = True
			self.bulk()
			if not self.search_canceled:
				self.allDone()
			self.bulk_search = False
			if self.search_canceled:
				self.search_canceled = False
			return
		
		##Access user search
		data = self.tbox.get()
		if not data: #check for empty search
			return
		self.last_search = data #stored for cancel search functionality
		
		## Pre Search Display configurations
		self.compare_lists_button.config(state=tk.DISABLED)
		self.bulk_search_button.config(state=tk.DISABLED)
		#self.chk.config(state=tk.DISABLED)
		self.search_button.config(state=tk.DISABLED)
		#self.save_txt_button.config(state=tk.DISABLED)
		#self.save_csv_button.config(state=tk.DISABLED)
		self.clear_button.config(state=tk.DISABLED)
		self.searchLabel.config(text="Searching Now!")
		self.cancel_button.config(state=tk.ACTIVE)
		self.clear_output()
		
		self.getBids(data) #main method call to print data
		
		## Post Search Display configurations
		#if self.chk_state.get():
			## Checks initially if save button was checked to disable the post run save as csv button
			#self.save_csv_button.config(state=tk.DISABLED)
		#else:
			#self.save_csv_button.config(state=tk.ACTIVE)
		if self.search_canceled:
			#self.save_txt_button.config(state = tk.DISABLED)
			#self.save_csv_button.config(state = tk.DISABLED)
			self.searchLabel.config(text ="Search Cancelled!")
		else:
			self.searchLabel.config(text="Done!")
			self.save_txt()
			#self.save_txt_button.config(state=tk.ACTIVE)
		self.clear_button.config(state=tk.ACTIVE)
		#self.chk.config(state=tk.ACTIVE)
		## if csv checked, add data to csv_list to ensure that csv will never be eliminated
		if self.chk_state.get() and not self.search_canceled :
			self.csv_list.append(data)
		self.search_canceled = False	
		self.compare_lists_button.config(state=tk.ACTIVE)
		self.bulk_search_button.config(state=tk.ACTIVE)
		self.cancel_button.config(state = tk.DISABLED)
	
	def clearAll(self):
		"Resets text windows to search for another item"
		self.clearExceptSearch() ## Calls different method to mostly clear
		self.tbox.delete(0,tk.END)
		self.saveLabel.config(text= "")
		self.search_canceled = True
	
	def clearAll_2(self):
		"Resets text windows to search for another item. Used for bulk search label functionality"
		self.clearExceptSearch_2() ## Calls different method to mostly clear
		self.tbox.delete(0,tk.END)
		self.saveLabel.config(text= "")
		
	def clearExceptSearch(self):
		"Resets except for text currently in search bar. Used for cancel search functionality"
		self.txt.delete(1.0,tk.END)
		self.tbox.focus()
		self.search_button.config(state=tk.ACTIVE)
		self.searchLabel.config(text="All Cleared!")
		self.saveLabel.config(text="")
		#self.save_txt_button.config(state=tk.DISABLED)
		#self.save_csv_button.config(state=tk.DISABLED)
		self.clear_button.config(state=tk.DISABLED)
		self.master.bind('<Return>', self.entr)
	
	def clearExceptSearch_2(self):
		"Resets except for text currently in search bar. Used for cancel search functionality and bulk search label functionality"
		self.txt.delete(1.0,tk.END)
		self.tbox.focus()
		self.search_button.config(state=tk.ACTIVE)
		search_label_string = "Searching! "+ str(self.bulk_count) + "/" + str(self.bulk_num_items)
		self.searchLabel.config(text=search_label_string)
		self.saveLabel.config(text="")
		#self.save_txt_button.config(state=tk.DISABLED)
		#self.save_csv_button.config(state=tk.DISABLED)
		self.clear_button.config(state=tk.DISABLED)
		self.master.bind('<Return>', self.entr)
		
	def save_txt(self):
		"Outputs the formatted text from the text window to a .txt file"
		if self.bulk_search and self.search_canceled:
			return
		if not self.save_txt_item_found:
			#self.save_txt_button.config(state=tk.DISABLED)
			return
		self.save_txt_item_found= False
		cur_inp = self.txt.get("1.0", tk.END) #text to save
		today = date.today() 
		date_str = today.strftime("%Y-%m-%d_Bid_Tabulations")
		parent_folder = "BTT Outputs"
		folder_path = os.path.join(self.parent_dir, parent_folder, date_str) #folder name
		self.shortest_abbreviated = self.shortest_abbreviated.replace('/', '_').replace('\\', '_').replace('-','_').replace('"', ' ').replace("'", ' ').replace('(', '').replace(')', '') #remove all bad characters that could cause .csv creation to fail
		## File path
		txt_file_path = os.path.join(folder_path, self.paynum + '_' + self.shortest_abbreviated + '.txt')
		
		## Make master folder if it doesn't exist
		if not os.path.exists(parent_folder):
			os.makedirs(parent_folder)
		
		## Make folder if it doesn't exist
		if not os.path.exists(folder_path):
			os.makedirs(folder_path)
			
		## Write the .txt file
		with open(txt_file_path, 'w') as txtfile:
			txtfile.write(cur_inp)
		
		## Display configurations
		self.saveLabel.config(text="Output Saved!")
		#self.save_txt_button.config(state=tk.DISABLED)
		
	def save_csv(self):
		"Outputs the formatted text from the text window to a .csv file"
		if self.bulk_search and self.search_canceled:
			return
		cur_inp = self.txt.get("1.0", tk.END) #text to save
		today = date.today() 
		date_str = today.strftime("%Y-%m-%d_Bid_Tabulations")
		parent_folder = "BTT Outputs"
		folder_path = os.path.join(self.parent_dir, parent_folder, date_str) #folder name
		self.shortest_abbreviated = self.shortest_abbreviated.replace('/', '_').replace('\\', '_').replace('-','_').replace('"', ' ').replace("'", ' ').replace('(', '').replace(')', '')#remove all bad characters that could cause .csv creation to fail
		self.csv_file_path = os.path.join(folder_path, self.paynum + '_' + self.shortest_abbreviated + '.csv')
		## File paths
		file_path = self.csv_file_path
		
		if os.path.exists(file_path):
			#self.save_csv_button.config(state=tk.DISABLED)
			return
		## Make master folder if it doesn't exist
		if not os.path.exists(parent_folder):
			os.makedirs(parent_folder)
		
		## Make folder if it doesn't exist
		if not os.path.exists(folder_path):
			os.makedirs(folder_path)
		
		## Remove formatting lines so .csv just has data
		lines = cur_inp.split("\n") #splits lines 
		data_lines = [line for line in lines if line.strip() and "Bid Tabulations" not in line] #removes formatting lines
		
		#Cases for removing lines based on how many text lines are printed before data
		if self.case1:
			num_remove =1
		if self.case2:
			num_remove =2
		if self.case3:
			num_remove =3
		
		##Remove lines at the start
		if len(data_lines)> num_remove:
			data_lines = data_lines[num_remove:]
		
		## Remove first four lines about item name and save the header
		filtered_lines = [data_lines[0]]
		
		## Remove subsequent instances of lines starting with "Contract"
		contract_count = 0
		for line in data_lines[1:]:
			if "Contract" not in line: #removes the rest of the header lines
				## Append each element to the filtered lines list
				filtered_lines.append(line)
		
		## Process the data into columns and rows so csv saves correctly
		rows =[]
		row =[]
		for line in filtered_lines:
			words = line.split()
			for word in words:
				if len(row)<8:
					row.append(word.replace(" ", ""))
				else:
					rows.append(row)
					row = [word.replace(" ", "")]
		rows.append(row)
		
		
		## Write the .csv file
		with open(file_path, 'w', newline = '') as csvfile:
			csv_writer = csv.writer(csvfile)
			for row in rows:
				csv_writer.writerow(row)
		
		## add to csv_list so never deleted
		self.csv_list.append(self.paynum)

		## Display configurations
		self.saveLabel.config(text="Output Saved!")
		#self.save_csv_button.config(state=tk.DISABLED)

	def cancel_search_func(self):
		"Cancels the current search and clears the screen except for the current search bar"
		self.cancel_search = True
		self.search_canceled = True
		self.clearExceptSearch()
		self.cancel_button.config(state = tk.DISABLED)
		#self.save_txt_button.config(state = tk.DISABLED)
		
		if self.paynum not in self.csv_list and os.path.exists(self.csv_file_path):
		## csv_list contains completed csv file names, thus self.paynum csv file not previously completed
			os.remove(self.csv_file_path) #remove it
			self.csv_file_path = ""
			
		self.clearExceptSearch()

	## Main search method
	def getBids(self, piNum):
		"""This function checks each excel file in the subfolders to see if the user input value matches any
		row in the first column.  A match result in the entire row being processed, filtered, and then output
		to the text window"""
		
		main_folder_path = self.source_data_file_path
		files = [fn for fn in glob.glob(os.path.join(main_folder_path, '*', '*.xls*'), recursive=True) if os.path.basename(os.path.dirname(fn)).startswith('20') and os.path.basename(os.path.dirname(fn)).endswith(' Bid Tabulations')]
		
		## Column names string
		# colNames = ['Contract', 'Quantity', 'Avg Bid', 'Low Price', 'High Price', 'Ttl # of PIs', 'Contract Total']
		colNames = ['Contract', 'Quantity', 'Avg Bid', 'Low Price', 'High Price', 'Ttl # of PIs', '# of Bidders', 'Contract Total']
		
		## Converts the user input to uppercase
		self.paynum = piNum.upper()
		# self.paynum = "20200100"		## EARTH EXCAVATION - for troubleshooting
		# self.paynum = "JT160327"		## FIBER OPTIC CABLE, SINGLE MODE, ARMORED, 864 FIBERS - for troubleshooting
		# 89502350
		
		## functionality for auto scroll enabled/disabled
		auto_scroll_enabled = self.auto_scroll_var.get()
		if not auto_scroll_enabled:
			self.toggle_auto_scroll()
		
		## Prints the name of the item being searched
		namefound = False
		found_in_1 = False
		# workbook1 = xlrd.open_workbook(r"E:\OneDrive - J.A. Watts Inc\Tollway General\Bid Tabulations\Source Data DO NOT EDIT\Tollway Coded Pay Items 11_07_2022.xlsx") #file path of name sheet to search
		workbook1 = xlrd.open_workbook(self.tollway_coded_file_path) #file path of name sheet to search
		sheet = workbook1.sheet_by_name("Sheet1")
		
		#iterate through all rows in column 3 of the sheet
		for row_idx in range(sheet.nrows):
			value = sheet.cell_value(row_idx, 2)
			if value == self.paynum:
			#grab the value directly to the right of the item code
				adjacent_cell = sheet.cell(row_idx, 3)
				namefound = True
				found_in_1 = True
				break
		if not namefound: #if not found in first sheet, iterate through second sheet
			workbook2 = xlrd.open_workbook(self.IDOT_coded_file_path)
			sheet2 = workbook2.sheet_by_name("Sheet1")
			#this time all rows in column 1
			for row_idx in range(sheet2.nrows):
				value = sheet2.cell_value(row_idx, 0)
				if value == self.paynum:
					#on alternate sheet, also grab value directly right of item code
					adjacent_cell = sheet2.cell(row_idx, 1)
					namefound = True
					break
					
						
		#insert at the beginning 
		if not namefound:
			self.txt.insert("1.0", self.paynum + " name not found in Tollway Coded Pay Items 11_07_2022 or 20230804HWYCodedPayItems" + "\n\n")
			self.case1 = True
		else:
			if found_in_1:
				self.txt.insert("1.0", "Item name: " + adjacent_cell.value+ "\n(Found in Tollway Coded Pay Items 11_07_2022)" + "\n\n")
				self.case2 = True
			else:
				self.txt.insert("1.0", "Item name: " + adjacent_cell.value+ "\n(found in 20230804HWYCodedPayItems)" + "\nONLY SEARCHING TOLLWAY USAGE OF " + adjacent_cell.value +"\n\n")
				self.case3 = True
		
		## Variable Declarations for search
		year = "Nothing"
		first = False
		firstTimeThrough = True #used to make sure header is only printed once upon the first entry found

		## For each excel file found in the subfolders...
		for f in files:
			##At each file search, cancel search if button pressed
			if self.cancel_search:
				break
				
			## Open Workbook
			wb = xlrd.open_workbook(f)
			sheet = wb.sheet_by_index(0)
			
			## Break down the file path to determine the folder (aka year) the file is in
			fname = os.path.split(f)
			yr = os.path.split(fname[0])
			
			# print(fname)	## '.\\YEAR Bid Tabulations', 'CONTRACT Bid Tabulations.xlsx'
			# print(yr)		## '.', 'YEAR Bid Tabulations'
			
			## If the year in the file path is different than current, print the new year 
			## and update the current year (so the year is only printed once per folder)
			if year != yr[1]:
				year = yr[1]
				if first == True:
					self.txt.insert("end"," No PI usage found.\n")
				self.txt.insert("end","\n"+year+":")
				if self.auto_scroll:
					self.txt.see(tk.END)
				first = True
				# self.writeLine()
			
			## Further breaks down file path to determine the file name (aka contract)
			name, ext = os.path.splitext(fname[1])
			
			## Removes "Bid Tabulations", etc. leaving only the contract
			contract = name.partition(' ')[0]
			
			## Gets the amount of the winning bid
			winner = self.winBid(sheet)
			
			## For each row...
			for i in range(sheet.nrows):
				
				## Check if the value in the first column matches the desired pay item...
				if sheet.cell_value(i, 0) == self.paynum:
					found = True
					self.save_txt_item_found = True
					## If this is the first contract to include the pay item for the year the header row is printed.
					if first == True:
						self.writeLine()
						first = False
					
					## Gets the data from that row (ignoring the first few columns)
					vals = list(filter(None, sheet.row_values(i)[4:]))
					
					## Creates a few variables
					avg = 0
					count = 0
					prices = []
					
					## Formats the row values
					for j in range(len(vals[:-1])):
						## Only includes the per unit bid price for each bidder
						if j % 2 == 0:
							prices.append(vals[j])
							avg = avg + vals[j]
							count += 1
							continue
					
					## Calculates the lowest, highest, and average per unit bid prices 
					lowPrc = min(prices)
					hiPrc = max(prices)
					avg = avg / count
					numBids = len(prices)
					
					## Creates a list of data to output, and then outputs it
					out = [contract, sheet.row_values(i)[3], round(avg,2), round(lowPrc,2), round(hiPrc,2), sheet.nrows-9, numBids, winner]
					self.txt.insert("end","{:>11}{:>7.0f}{:>12.2f}{:>12.2f}{:>12.2f}{:>11}{:>7}{:>13.3f}\n".format(contract, sheet.row_values(i)[3], round(avg,2), round(lowPrc,2), round(hiPrc,2), sheet.nrows-9, numBids, winner/1000000))
					##	Contract	Qty      	Avg     	 Low    	 High   TotPIs     Total(mil)
					##	{:>11}	{:>7.0f}	{:>12.2f}	{:>12.2f}	{:>12.2f}	{:>11}		{:>13.3f}
					
					## Scroll to the end upon printing data 
					if self.auto_scroll:
						self.txt.see(tk.END)
					
					## Checks to see if the "Save?" checkbox was checked
					if self.chk_state.get() and self.paynum not in self.csv_list:
					## If so, .csv file will be created/saved
						writeOn = True
					else:
						writeOn = False
					
					if firstTimeThrough:
						## Still want to create file path if user decides to save after
						self.abbreviated_names.append(sheet.cell_value(i, 1))
						self.shortest_abbreviated = min(self.abbreviated_names, key = len)
						self.shortest_abbreviated = self.shortest_abbreviated.replace('/', '_').replace('\\', '_').replace('-','_').replace('"', ' ').replace("'", ' ').replace('(', '').replace(')', '')#remove all bad characters that could cause .csv creation to fail
						self.current_shortest_name = self.shortest_abbreviated
						## Date information for folder name
						today = date.today()
						date_str = today.strftime("%Y-%m-%d_Bid_Tabulations")
						parent_folder = "BTT Outputs"
						folder_path = os.path.join(self.parent_dir, parent_folder, date_str)  # folder name
						self.csv_file_path = os.path.join(folder_path, self.paynum + '_' + self.shortest_abbreviated + '.csv')
					
					## If Save was checked, the column/header names are written to the .csv file only once
					if writeOn and firstTimeThrough:
						## Make parent folder if it doesn't exist
						if not os.path.exists(parent_folder):
							os.makedirs(parent_folder)

						## Create a folder if it doesn't already exits
						if not os.path.exists(folder_path):
							os.makedirs(folder_path)
			
						## Write to file 
						with open(self.csv_file_path, 'a', newline='') as csvfile:
							outWriter = csv.writer(csvfile) #, delimiter=' ')
							outWriter.writerow(colNames)
						## Change flag so headers aren't printed again
						firstTimeThrough = False
					
					## At each iteration write the information to the file
					if writeOn:
						self.abbreviated_names.append(sheet.cell_value(i, 1))
						self.shortest_abbreviated = min(self.abbreviated_names, key = len)
						self.shortest_abbreviated = self.shortest_abbreviated.replace('/', '_').replace('\\', '_').replace('-','_').replace('"', ' ').replace("'", ' ').replace('(', '').replace(')', '')#remove all bad characters that could cause .csv creation to fail
						## If the shortest name has changed, update the CSV file name
						if self.shortest_abbreviated != self.current_shortest_name and os.path.exists(self.csv_file_path):
							new_csv_file_path = os.path.join(folder_path, self.paynum + '_' + self.shortest_abbreviated + '.csv')
							os.rename(self.csv_file_path, new_csv_file_path)
							self.current_shortest_name = self.shortest_abbreviated
							self.csv_file_path = new_csv_file_path
						## Writes to csv
						with open(self.csv_file_path, 'a', newline='') as csvfile:
							outWriter = csv.writer(csvfile)
							outWriter.writerow(out)
					else:
						self.abbreviated_names.append(sheet.cell_value(i, 1))
						self.shortest_abbreviated = min(self.abbreviated_names, key = len)
						if self.shortest_abbreviated != self.current_shortest_name and os.path.exists(self.csv_file_path):
							new_csv_file_path = os.path.join(folder_path, self.paynum + '_' + self.shortest_abbreviated + '.csv')
							self.current_shortest_name = self.shortest_abbreviated
							self.csv_file_path = new_csv_file_path
			
			## Updates the GUI text window
			self.txt.update()
			
		
		## Once all files in the folders have been checked, this runs to see if the last year had any usages, and if not it writes a message saying so.
		else:
			if first == True:
				self.txt.insert("end"," No PI usage found.")
				if self.auto_scroll:
					self.txt.see(tk.END)
		if not self.cancel_search:
			self.csv_list.append(self.paynum)			
		## Formatting		
		self.txt.insert("end","\n")
		if self.auto_scroll:
			self.txt.see(tk.END)
			
		self.cancel_search = False		

	## Helper Methods
	def allDone(self):
		"Fun display for when a bulk search has completed"
		tabs = "\t\t\t\t\t"
		newLs = "\n\n\n\n\n"
		self.txt.insert(1.0,newLs+newLs+newLs)
		self.txt.insert(19.0,tabs+" _____   ____  _   _ ______ _ \n")
		self.txt.insert(20.0,tabs+"|  __ \ / __ \| \ | |  ____| |\n")
		self.txt.insert(21.0,tabs+"| |  | | |  | |  \| | |__  | |\n")
		self.txt.insert(22.0,tabs+"| |  | | |  | | . ` |  __| | |\n")
		self.txt.insert(23.0,tabs+"| |__| | |__| | |\  | |____|_|\n")
		self.txt.insert(24.0,tabs+"|_____/ \____/|_| \_|______(_)\n")	

	def display_image(self, image_path, width, height):
		"Functionality for displaying JWI logo"
		image=Image.open(image_path)
		image=image.resize((width,height),Image.Resampling.LANCZOS)
		photo = ImageTk.PhotoImage(image)
		self.image_label.configure(image=photo)
		self.image_label.image = photo

	def winBid(self,sht):
		""" Locates the winning bid value by finding the data header row and moving over and up from there.
			This is necessary due to various inconsistencies in the information at the top of the sheet"""
		for r in range(19):
			if sht.cell_value(r, 0) == "Item ":
				return sht.cell_value(r-1, 5)
				
	def clear_output(self):
		"Clears the screen"
		self.txt.delete(1.0,tk.END)
				
	def writeLine(self):
		"Writes the header to the GUI text window"
		# self.txt.insert("end","\n{:>11}{:>7}{:>12}{:>12}{:>12}{:>11}{:>15}\n".format('Contract', 'Qty', 'Avg($)', 'Low($)', 'High($)', 'TotPIs', 'Total(mil$)'))
		self.txt.insert("end","\n{:>11}{:>7}{:>12}{:>12}{:>12}{:>11}{:>8}{:>14}\n".format('Contract', 'Qty', 'Avg($)', 'Low($)', 'High($)', 'TotPIs', '#Bids', 'Total(mil$)'))
	
	
## Runs the program!
root = Tk()
my_gui = GUIwindow(root)
root.mainloop()






